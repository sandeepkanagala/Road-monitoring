from flask import Flask, request, jsonify
from flask import send_from_directory, send_file
from flask_cors import CORS
from datetime import datetime
import json
import os
import io

# Excel library
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

app = Flask(__name__)
CORS(app)

DATA_FILE = 'road_data.json'



# ---------------- LOAD/SAVE FILE ---------------- #

DEVICES_FILE = 'devices.json'

def load_data():
    """Safely load saved data."""
    if not os.path.exists(DATA_FILE):
        return []
    try:
        with open(DATA_FILE, 'r') as f:
            contents = f.read().strip()
            if not contents:
                return []
            return json.loads(contents)
    except (json.JSONDecodeError, OSError) as exc:
        print(f"⚠️ Error reading {DATA_FILE}: {exc}. Returning blank list.")
        return []


def save_data(data):
    with open(DATA_FILE, 'w') as f:
        json.dump(data, f, indent=2)





# ---------------- API: MANAGED DEVICES ---------------- #







# ---------------- UPLOAD DIRECTORIES ---------------- #

IMAGE_UPLOAD_DIR = "uploads/images"
VIDEO_UPLOAD_DIR = "uploads/videos"

# base directories exist; per-device subfolders will be created as needed
os.makedirs(IMAGE_UPLOAD_DIR, exist_ok=True)
os.makedirs(VIDEO_UPLOAD_DIR, exist_ok=True)



# ---------------- API: RECEIVE SENSOR DATA ---------------- #

@app.route('/api/sensor-data', methods=['POST'])
def receive_sensor_data():
    data = request.json or {}

    data['timestamp'] = datetime.now().isoformat()

    # Attach device id if provided by client
    device_id = data.get('device_id')
    if device_id:
        data['device_id'] = str(device_id)

    # Accept both Android formats:
    # Android may send accel_x or x
    data['accel_x'] = data.get('accel_x') or data.get('x', 0)
    data['accel_y'] = data.get('accel_y') or data.get('y', 0)
    data['accel_z'] = data.get('accel_z') or data.get('z', 0)

    # Magnitude
    data['accel_magnitude'] = (
        data.get('accel_magnitude')
        or data.get('accelerometer', 0)
    )

    # Save
    all_data = load_data()
    all_data.append(data)

    if len(all_data) > 1000:
        all_data = all_data[-1000:]

    save_data(all_data)

    print(f"✓ Received - Lat: {data.get('latitude')}, Lon: {data.get('longitude')}, "
          f"X={data['accel_x']} Y={data['accel_y']} Z={data['accel_z']} "
          f"Mag={data['accel_magnitude']}")

    return jsonify({'status': 'success'}), 200


# ---------------- API: GET LATEST DATA ---------------- #

@app.route('/api/get-latest', methods=['GET'])
def get_latest():
    data = load_data()
    device_id = request.args.get('device_id')
    if device_id:
        filtered = [d for d in data if str(d.get('device_id')) == str(device_id)]
        return jsonify(filtered[-1000:] if filtered else []), 200
    return jsonify(data[-1000:] if data else []), 200


# ---------------- API: ROAD QUALITY ---------------- #

@app.route('/api/road-quality', methods=['GET'])
def calculate_road_quality():
    data = load_data()
    device_id = request.args.get('device_id')
    if device_id:
        data = [d for d in data if str(d.get('device_id')) == str(device_id)]
    if not data:
        return jsonify({'quality': 'No data'}), 200

    magnitudes = [abs(item.get('accel_magnitude', 0)) for item in data]
    avg_accel = sum(magnitudes) / len(magnitudes)

    if avg_accel < 2:
        quality = "Excellent"
    elif avg_accel < 5:
        quality = "Good"
    elif avg_accel < 10:
        quality = "Fair"
    else:
        quality = "Poor"

    return jsonify({
        'quality': quality,
        'avg_vibration': round(avg_accel, 2),
        'total_readings': len(data)
    }), 200


@app.route('/api/devices', methods=['GET'])
def list_devices():
    """Return a list of known device IDs based on data and upload folders."""
    data = load_data()
    devices = set()
    for d in data:
        if d.get('device_id'):
            devices.add(str(d.get('device_id')))

    # Also include subfolders from uploads
    try:
        for name in os.listdir(IMAGE_UPLOAD_DIR):
            p = os.path.join(IMAGE_UPLOAD_DIR, name)
            if os.path.isdir(p):
                devices.add(name)
    except Exception:
        pass

    try:
        for name in os.listdir(VIDEO_UPLOAD_DIR):
            p = os.path.join(VIDEO_UPLOAD_DIR, name)
            if os.path.isdir(p):
                devices.add(name)
    except Exception:
        pass

    devices_list = sorted(list(devices))
    return jsonify({'devices': devices_list}), 200


# ---------------- API: CLEAR DATA ---------------- #

@app.route('/api/clear-data', methods=['POST'])
def clear_data():
    save_data([])
    return jsonify({'status': 'success', 'message': 'Data cleared'}), 200


# ---------------- EXPORT TO EXCEL (XYZ + Magnitude) ---------------- #

@app.route('/api/export-excel', methods=['GET'])
def export_excel():
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'openpyxl not installed'}), 500

    data = load_data()
    # optional device filter
    device_id = request.args.get('device_id')
    if device_id:
        data = [d for d in data if str(d.get('device_id')) == str(device_id)]
    if not data:
        return jsonify({'error': 'No data to export'}), 404

    wb = Workbook()
    ws = wb.active
    ws.title = "Road Monitoring Data"

    headers = [
        'Device ID',
        'Timestamp', 'Date', 'Time',
        'Latitude', 'Longitude',
        'Accel X', 'Accel Y', 'Accel Z', 'Magnitude'
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="667eea", end_color="764ba2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for item in data:
        timestamp = item.get('timestamp', '')
        try:
            dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
            date_str = dt.strftime("%Y-%m-%d")
            time_str = dt.strftime("%H:%M:%S")
        except:
            date_str = ""
            time_str = ""

        ws.append([
            item.get('device_id', ''),
            timestamp,
            date_str,
            time_str,
            item.get('latitude', 0),
            item.get('longitude', 0),
            item.get('accel_x', 0),
            item.get('accel_y', 0),
            item.get('accel_z', 0),
            item.get('accel_magnitude', 0)
        ])

    for column in ws.columns:
        length = max(len(str(cell.value)) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"road_monitoring_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ---------------- API: UPLOAD IMAGE ---------------- #

@app.route('/api/upload-image', methods=['POST'])
def upload_image():
    if 'image' not in request.files:
        return jsonify({'error': 'No image file provided'}), 400

    image = request.files['image']
    if image.filename == '':
        return jsonify({'error': 'Empty filename'}), 400

    filename = f"image_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
    # optionally save under a device subfolder
    device_id = request.form.get('device_id') or request.args.get('device_id')
    if device_id:
        device_dir = os.path.join(IMAGE_UPLOAD_DIR, str(device_id))
        os.makedirs(device_dir, exist_ok=True)
        filepath = os.path.join(device_dir, filename)
    else:
        filepath = os.path.join(IMAGE_UPLOAD_DIR, filename)
    image.save(filepath)

    print(f"📸 Image saved: {filepath}")

    return jsonify({
        'status': 'success',
        'file': filename,
        'device_id': device_id
    }), 200


# ---------------- API: UPLOAD VIDEO ---------------- #

@app.route('/api/upload-video', methods=['POST'])
def upload_video():
    if 'video' not in request.files:
        return jsonify({'error': 'No video file provided'}), 400

    video = request.files['video']
    if video.filename == '':
        return jsonify({'error': 'Empty filename'}), 400

    filename = f"video_{datetime.now().strftime('%Y%m%d_%H%M%S')}.mp4"
    # optionally save under a device subfolder
    device_id = request.form.get('device_id') or request.args.get('device_id')
    if device_id:
        device_dir = os.path.join(VIDEO_UPLOAD_DIR, str(device_id))
        os.makedirs(device_dir, exist_ok=True)
        filepath = os.path.join(device_dir, filename)
    else:
        filepath = os.path.join(VIDEO_UPLOAD_DIR, filename)
    video.save(filepath)

    print(f"🎥 Video saved: {filepath}")

    return jsonify({
        'status': 'success',
        'file': filename,
        'device_id': device_id
    }), 200


# ---------------- API: GET IMAGES LIST ---------------- #

@app.route('/api/get-images', methods=['GET'])
def get_images():
    """Get list of all uploaded images."""
    try:
        if not os.path.exists(IMAGE_UPLOAD_DIR):
            return jsonify({'images': []}), 200

        device_id = request.args.get('device_id')
        images = []

        # If a device_id is provided, list only that folder
        if device_id:
            target = os.path.join(IMAGE_UPLOAD_DIR, device_id)
            if not os.path.exists(target):
                return jsonify({'images': []}), 200
            for filename in os.listdir(target):
                if filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp')):
                    filepath = os.path.join(target, filename)
                    stat = os.stat(filepath)
                    images.append({
                        'filename': filename,
                        'device_id': device_id,
                        'url': f'/api/image/{device_id}/{filename}',
                        'size': stat.st_size,
                        'modified': datetime.fromtimestamp(stat.st_mtime).isoformat()
                    })
        else:
            # include files in root images folder
            for entry in os.listdir(IMAGE_UPLOAD_DIR):
                full = os.path.join(IMAGE_UPLOAD_DIR, entry)
                if os.path.isfile(full) and entry.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp')):
                    stat = os.stat(full)
                    images.append({
                        'filename': entry,
                        'device_id': None,
                        'url': f'/api/image/{entry}',
                        'size': stat.st_size,
                        'modified': datetime.fromtimestamp(stat.st_mtime).isoformat()
                    })
                elif os.path.isdir(full):
                    # files inside device subfolders
                    device = entry
                    for filename in os.listdir(full):
                        if filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp')):
                            filepath = os.path.join(full, filename)
                            stat = os.stat(filepath)
                            images.append({
                                'filename': filename,
                                'device_id': device,
                                'url': f'/api/image/{device}/{filename}',
                                'size': stat.st_size,
                                'modified': datetime.fromtimestamp(stat.st_mtime).isoformat()
                            })

        # Sort by modified time, newest first
        images.sort(key=lambda x: x['modified'], reverse=True)
        return jsonify({'images': images}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ---------------- API: GET VIDEOS LIST ---------------- #

@app.route('/api/get-videos', methods=['GET'])
def get_videos():
    """Get list of all uploaded videos."""
    try:
        if not os.path.exists(VIDEO_UPLOAD_DIR):
            return jsonify({'videos': []}), 200

        device_id = request.args.get('device_id')
        videos = []

        if device_id:
            target = os.path.join(VIDEO_UPLOAD_DIR, device_id)
            if not os.path.exists(target):
                return jsonify({'videos': []}), 200
            for filename in os.listdir(target):
                if filename.lower().endswith(('.mp4', '.avi', '.mov', '.mkv', '.webm')):
                    filepath = os.path.join(target, filename)
                    stat = os.stat(filepath)
                    videos.append({
                        'filename': filename,
                        'device_id': device_id,
                        'url': f'/api/video/{device_id}/{filename}',
                        'size': stat.st_size,
                        'modified': datetime.fromtimestamp(stat.st_mtime).isoformat()
                    })
        else:
            for entry in os.listdir(VIDEO_UPLOAD_DIR):
                full = os.path.join(VIDEO_UPLOAD_DIR, entry)
                if os.path.isfile(full) and entry.lower().endswith(('.mp4', '.avi', '.mov', '.mkv', '.webm')):
                    stat = os.stat(full)
                    videos.append({
                        'filename': entry,
                        'device_id': None,
                        'url': f'/api/video/{entry}',
                        'size': stat.st_size,
                        'modified': datetime.fromtimestamp(stat.st_mtime).isoformat()
                    })
                elif os.path.isdir(full):
                    device = entry
                    for filename in os.listdir(full):
                        if filename.lower().endswith(('.mp4', '.avi', '.mov', '.mkv', '.webm')):
                            filepath = os.path.join(full, filename)
                            stat = os.stat(filepath)
                            videos.append({
                                'filename': filename,
                                'device_id': device,
                                'url': f'/api/video/{device}/{filename}',
                                'size': stat.st_size,
                                'modified': datetime.fromtimestamp(stat.st_mtime).isoformat()
                            })

        videos.sort(key=lambda x: x['modified'], reverse=True)
        return jsonify({'videos': videos}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ---------------- API: SERVE IMAGE ---------------- #

@app.route('/api/image/<path:filepath>', methods=['GET'])
def serve_image(filepath):
    """Serve an uploaded image file. Supports optional device subfolder: /api/image/<device_id>/<filename>"""
    parts = filepath.split('/', 1)
    if len(parts) == 1:
        return send_from_directory(IMAGE_UPLOAD_DIR, parts[0])
    device, filename = parts[0], parts[1]
    device_dir = os.path.join(IMAGE_UPLOAD_DIR, device)
    return send_from_directory(device_dir, filename)


# ---------------- API: SERVE VIDEO ---------------- #

@app.route('/api/video/<path:filepath>', methods=['GET'])
def serve_video(filepath):
    """Serve an uploaded video file. Supports optional device subfolder: /api/video/<device_id>/<filename>"""
    parts = filepath.split('/', 1)
    if len(parts) == 1:
        return send_from_directory(VIDEO_UPLOAD_DIR, parts[0])
    device, filename = parts[0], parts[1]
    device_dir = os.path.join(VIDEO_UPLOAD_DIR, device)
    return send_from_directory(device_dir, filename)


# ---------------- DASHBOARD / STATIC ROUTES ---------------- #

@app.route('/')
def index():
    return send_from_directory('.', 'dashboard.html')


@app.route('/styles.css')
def styles():
    return send_from_directory('.', 'styles.css')


@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'Server running'}), 200


# ---------------- SERVER START ---------------- #

if __name__ == '__main__':
    print("🚀 Road Monitoring Server Started")
    print("📍 Dashboard: http://localhost:5000")
    print("📱 POST URL: http://YOUR_PC_IP:5000/api/sensor-data")
    app.run(debug=True, host='0.0.0.0', port=5000)
